import subprocess
import time
import pyautogui
import pyperclip
import os
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import yaml
from settings import file, anotherFile, q, a, open_llm_script_template, focus_resize_script_template, copy_visible_text_script_template
from types import SimpleNamespace

def to_namespace(d): return SimpleNamespace(**{k: to_namespace(v) if isinstance(v, dict) else v for k, v in d.items()})
with open(file, "r") as f: raw_config = yaml.safe_load(f)
def copy_securely(text): subprocess.run("pbcopy", text=True, input=text)
    
config = to_namespace(raw_config)
click_coords = tuple(config.click_coords)
response_file_path = config.response_file_path
custom_delimiter = config.custom_delimiter
excel_file = config.excel_file
refresh_config = config.refresh_chat
llm = config.llm
browser_name = config.browser.name
llm_url = llm.url

copy_securely("")

def refresh_and_restart_chat():
    pyautogui.hotkey(*refresh_config.reload_hotkey)
    time.sleep(4)

    pyautogui.click(*refresh_config.click_reload_coords)
    time.sleep(0.5)

    for _ in range(refresh_config.chatbox_tab_count_before_new_chat):
        pyautogui.press('tab')
        time.sleep(0.1)

    pyautogui.press('return')
    time.sleep(1.5)

    pyautogui.click(*refresh_config.new_chat_click_coords)
    time.sleep(0.5)

    pyautogui.press('down', presses=refresh_config.down_presses, interval=0.2)
    pyautogui.press('return')
    time.sleep(0.5)

    pyautogui.press('tab')
    pyautogui.press('return')
    time.sleep(1)

    pyautogui.click(*refresh_config.click_reload_coords)
    time.sleep(0.5)

    for _ in range(refresh_config.final_tab_count_to_focus_input):
        pyautogui.press('tab')
        time.sleep(0.1)

    time.sleep(1)
    pyautogui.press('return')
    
def force_quit_browser():
    browser_name = config.browser.name
    try:
        subprocess.run(["pkill", "-f", browser_name], check=True)
        print(f"🛑 {browser_name} force-quit successfully.")
    except subprocess.CalledProcessError:
        print(f"⚠️ {browser_name} not running or already closed.")

def open_llm_in_browser():
    applescript = open_llm_script_template.format(
        browser_name=config.browser.name,
        llm_url=config.llm.url
    )
    subprocess.run(["osascript", "-e", applescript])
    print(f"🌐 LLM page '{config.llm.url}' requested in {config.browser.name}.")


def focus_and_resize_window():
    win_width, win_height = config.window.size
    win_x, win_y = config.window.position

    applescript = focus_resize_script_template.format(
        win_width=win_width,
        win_height=win_height,
        win_x=win_x,
        win_y=win_y
    )

    subprocess.run(["osascript", "-e", applescript])
    time.sleep(0.5)

def copy_visible_text():
    focus_and_resize_window()
    time.sleep(config.keyboard_delays.before_copy_click_wait)

    pyautogui.click(*click_coords)

    applescript = copy_visible_text_script_template.format(
        copy_delay=config.keyboard_delays.copy_delay
    )

    subprocess.run(["osascript", "-e", applescript])
    time.sleep(config.keyboard_delays.after_copy_wait)

    return pyperclip.paste()

def wait_for_llm_ui():
    ui_ready_text = config.llm.ui_ready_text

    while True:
        screen_text = copy_visible_text()
        if ui_ready_text in screen_text:
            print(f"✅ {config.llm.name} chat interface is active.")
            pyautogui.press('tab')  # Focus input box
            return True
        print(f"🔄 Waiting for {config.llm.name} to load...")
        time.sleep(1)

def type_into_chatbox(text):
    prompt_header = config.llm.prompt_header
    full_message = prompt_header + text

    # Save the full message to a file
    with open(anotherFile, "w", encoding="utf-8") as f:
        f.write(full_message)

    # Copy to clipboard
    copy_securely(full_message)

    # Get AppleScript from config
    apple_script = config.keyboard_macros.submit_clipboard_apple_script
    subprocess.run(["osascript", "-e", apple_script])

    print(f"📋 Pasted and sent to {config.llm.name}.")

def wait_for_llm_response():
    print(f"⏳ Waiting for {config.llm.name}'s final response...")

    while True:
        current_content = copy_visible_text()

        if config.llm.output_end_signal in current_content:
            with open(response_file_path, "w", encoding="utf-8") as f:
                f.write(current_content)
            print(f"✅ Final {config.llm.name} response saved (signal detected).")
            return True

        if config.llm.server_busy_text in current_content:
            print("⚠️ Server busy message detected. Will retry same section.")
            return False

        time.sleep(10)
def preprocess_response(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        text = f.read()

    if os.path.exists(anotherFile):
        with open(anotherFile, "r", encoding="utf-8") as r:
            request_text = r.read().strip()
        text = text.replace(request_text, "")

    start_index = text.find(q)
    if start_index == -1:
        return []

    cleaned_text = text[start_index:]
    parts = cleaned_text.split(custom_delimiter)
    return [part.strip() for part in parts if part.strip()]

def get_note_text(folder, note_name):
    script_template = config.notes.script_template

    applescript = script_template.format(
        app_name=config.notes.app_name,
        account_name=config.notes.account_name,
        folder=folder,
        note_name=note_name
    )

    try:
        result = subprocess.run(
            ["osascript", "-e", applescript],
            capture_output=True,
            text=True,
            check=True
        )
        note_body = result.stdout.strip()

        if config.notes.extract_html:
            soup = BeautifulSoup(note_body, "html.parser")
            return soup.get_text(separator="\n")
        else:
            return note_body

    except subprocess.CalledProcessError as e:
        print(f"❌ Failed to read note from {config.notes.app_name}:", e.stderr)
        return None

# 🧼 Start session
force_quit_browser()
open_llm_in_browser()

if wait_for_llm_ui():
    print(f"🎯 Ready to interact with {config.llm.name}.")

    folder_name = config.note_input.folder_name
    note_title = config.note_input.note_title
    section_delimiter = config.note_input.section_delimiter

    note_content = get_note_text(folder_name, note_title)

    if note_content:
        sections = [s.strip() for s in note_content.split(section_delimiter) if s.strip()]
        print(f"📄 Total sections found: {len(sections)}")

        q_list, a_list = [], []
        idx = 0

        while idx < len(sections):
            selected_text = sections[idx]
            copy_securely("")

            if idx > 0:
                refresh_and_restart_chat()
                time.sleep(1)
                pyautogui.click(*click_coords)
                pyautogui.press('tab')

            print(f"📤 Sending section {idx + 1}...")
            type_into_chatbox(selected_text)

            success = wait_for_llm_response()
            if not success:
                print(f"🔁 Retrying section {idx + 1} due to failure...")
                time.sleep(config.retry.delay_seconds)
                continue

            qa_pairs = preprocess_response(response_file_path)

            for entry in qa_pairs:
                lines = entry.splitlines()
                current_q, current_a = None, None

                for line in lines:
                    if line.strip().startswith(q):
                        if current_q and current_a:
                            q_list.append(current_q)
                            a_list.append(current_a)
                            current_q, current_a = None, None
                        current_q = line.strip()[2:].strip()
                    elif line.strip().startswith(a):
                        current_a = line.strip()[2:].strip()
                    else:
                        if current_a is not None:
                            current_a += " " + line.strip()

                if current_q and current_a:
                    q_list.append(current_q)
                    a_list.append(current_a)

            idx += 1

        df_new = pd.DataFrame({'Question': q_list, 'Answer': a_list})

        if os.path.exists(excel_file):
            book = load_workbook(excel_file)
            sheet = book.active
            start_row = sheet.max_row

            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df_new.to_excel(writer, index=False, header=False, startrow=start_row)

            print(f"📎 Appended {len(df_new)} flashcards to '{excel_file}'")
        else:
            df_new.to_excel(excel_file, index=False)
            print(f"📁 Created new Excel file and exported {len(df_new)} flashcards to '{excel_file}'")
    else:
        print("⚠️ Could not retrieve note content.")

# 📱 Notify via iMessage
message = config.export.success_message
phone_number = config.export.phone_number
imessage_script = config.export.imessage_applescript.format(
    message=message,
    phone_number=phone_number
)
subprocess.run(["osascript", "-e", imessage_script])